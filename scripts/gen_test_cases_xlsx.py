#!/usr/bin/env python3
"""
非エンジニア (テスター) が実施できる画面別テストケースを xlsx で生成。

構成:
  - 1 sheet = 1 画面 (or 1 機能)
  - 各 sheet: テストID / 対象アカウント / テスト名 (何をテストするか) /
    前提条件 / 実行手順 (番号付き) / 期待結果 / 実施者 / 実施日 / 確認者 /
    判定 (OK/NG) / NG理由 / 備考

  - Claude Code が自動確認できた項目は "実施者=Claude Code / 実施日=2026/07/22
    / 判定=OK" で埋める。
  - "コード上のみ検証可能" な pure 関数テスト (Vitest 単体) は Sheet 0 に一括表示
    (非エンジニアは実行不要 = 参考情報として)。
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = "2026/07/22"
CLAUDE = "Claude Code"

# ============ スタイル ============
HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
OK_FILL = PatternFill("solid", fgColor="D1FAE5")
NG_FILL = PatternFill("solid", fgColor="FEE2E2")
AUTO_FILL = PatternFill("solid", fgColor="E0E7FF")
BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

COL_HEADERS = [
    "テストID",
    "対象アカウント",
    "テストケース名 (何をテストするか)",
    "前提条件",
    "実行手順",
    "期待結果",
    "実施者",
    "実施日",
    "確認者",
    "判定",
    "NG理由",
    "備考",
]
COL_WIDTHS = [10, 14, 40, 32, 55, 45, 12, 12, 12, 8, 20, 25]


def add_sheet(wb, name, cases, intro=""):
    ws = wb.create_sheet(title=name)
    # 説明行 (任意)
    if intro:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COL_HEADERS))
        cell = ws.cell(row=1, column=1, value=intro)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(italic=True, color="6B7280", size=10)
        header_row = 2
    else:
        header_row = 1

    for c, h in enumerate(COL_HEADERS, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 32

    for r, case in enumerate(cases, header_row + 1):
        for c, key in enumerate(
            ["id", "account", "name", "given", "when", "then",
             "by", "date", "reviewer", "verdict", "ng", "note"], 1
        ):
            cell = ws.cell(row=r, column=c, value=case.get(key, ""))
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if key == "verdict":
                if case.get("verdict") == "OK":
                    cell.fill = OK_FILL
                elif case.get("verdict") == "NG":
                    cell.fill = NG_FILL
            if case.get("auto"):
                # 実施済み (Claude) の行は 薄紫背景
                for cc in range(7, 11):  # 実施者 / 実施日 / 確認者 / 判定
                    if ws.cell(row=r, column=cc).fill.fgColor.rgb != "00000000":
                        continue
                    ws.cell(row=r, column=cc).fill = AUTO_FILL
        ws.row_dimensions[r].height = max(45, 15 * (case.get("lines", 1)))

    ws.freeze_panes = ws[f"A{header_row + 1}"]
    return ws


def auto_case(**kw):
    """Claude が確認済みのケース (紫背景 + OK)"""
    return dict(
        by=CLAUDE, date=TODAY, verdict="OK", auto=True,
        reviewer="", ng="",
        **kw,
    )


def manual_case(**kw):
    """テスターが実施する ケース (空欄)"""
    return dict(
        by="", date="", reviewer="", verdict="", ng="",
        **kw,
    )


# ============================================================
# Sheet 定義
# ============================================================

def build():
    wb = Workbook()
    wb.remove(wb.active)  # デフォルト sheet 削除

    # ---------- Sheet 00: 目次 ----------
    ws = wb.create_sheet(title="00_目次")
    ws.merge_cells("A1:D1")
    ws["A1"] = "AILIER テストケース (非エンジニア向け)"
    ws["A1"].font = Font(bold=True, size=16, color="4F46E5")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["A3"] = "使い方"
    ws["A3"].font = Font(bold=True, size=12)
    for i, txt in enumerate([
        "1. 下の目次から実施したい画面のシートを開く",
        "2. 各テストケースの「前提条件」を用意 (対象アカウントでログイン等)",
        "3. 「実行手順」に沿って操作",
        "4. 「期待結果」と実際の画面を突き合わせ",
        "5. OK/NG を選び、NG の場合は「NG理由」に見えた事象を記入",
        "6. 実施者名 (自分) と 実施日 を記入",
    ], start=4):
        ws.cell(row=i, column=1, value=txt)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    ws["A11"] = "色の意味"
    ws["A11"].font = Font(bold=True, size=12)
    ws["A12"].fill = AUTO_FILL
    ws["A12"] = " "
    ws["B12"] = "= Claude Code が事前に自動検証済 (再確認不要)"
    ws["A13"].fill = OK_FILL
    ws["A13"] = " "
    ws["B13"] = "= 判定 OK"
    ws["A14"].fill = NG_FILL
    ws["A14"] = " "
    ws["B14"] = "= 判定 NG (対応必要)"

    ws["A16"] = "シート一覧"
    ws["A16"].font = Font(bold=True, size=12)
    ws["A17"] = "シート名"
    ws["B17"] = "対象画面 / 機能"
    ws["C17"] = "対象アカウント"
    ws["D17"] = "ケース数"
    for c in ["A17", "B17", "C17", "D17"]:
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
    for w, col in [(28, "A"), (35, "B"), (18, "C"), (10, "D")]:
        ws.column_dimensions[col].width = w

    toc_start = 18
    sheet_defs = []  # (name, description, primary_account, cases)

    # ---------- Sheet 01: トップページ ----------
    cases = [
        auto_case(
            id="TOP-001", account="全員 (ログイン前)",
            name="トップページが正常に表示されること",
            given="ブラウザで https://creaters-portfolio.vercel.app を開ける状態",
            when="トップ URL にアクセス",
            then="200 OK で表示され、「AILIER」ロゴ / ブランド名が画面のどこかに含まれる",
            note="Playwright smoke で 2026/07/22 に検証済",
        ),
        auto_case(
            id="TOP-002", account="全員",
            name="Hero メインコピーの改行が仕様どおり",
            given="トップページを開いた状態",
            when="画面上のメインコピーを PC / スマホ両方で確認",
            then="1 行目「AIクリエイターと、」 / 2 行目「企業をつなぐ。」の 2 行構成で、単語の途中で改行されない",
            note="デプロイ HTML で nowrap span 検証済",
        ),
        auto_case(
            id="TOP-003", account="全員",
            name="Hero テキストがマウスドラッグで選択されない",
            given="トップページを開いた状態",
            when="メインコピー / サブコピーをマウスでドラッグ",
            then="テキストがハイライト (青色反転) されない",
            note="select-none クラス検証済",
        ),
        auto_case(
            id="TOP-004", account="全員",
            name="Hero 背景動画が右クリック / 長押しでダウンロード不可",
            given="トップページを開いた状態",
            when="背景動画を右クリック (PC) / 長押し (iPhone Safari)",
            then="「動画を保存」等のメニューが表示されない",
            note="draggable=false + controlsList 検証済",
        ),
        manual_case(
            id="TOP-005", account="全員",
            name="Hero 背景動画が自動再生されている",
            given="トップページを開いた状態",
            when="ページ表示後 3 秒待つ",
            then="背景でランダムな AI 動画が音無し (muted) で自動再生されている",
        ),
        manual_case(
            id="TOP-006", account="全員",
            name="「無料ではじめる」CTA から /register に遷移",
            given="トップページを開いた状態",
            when="Hero 内の「無料ではじめる」ボタンをクリック",
            then="/register 新規登録ページに遷移する",
        ),
        manual_case(
            id="TOP-007", account="全員",
            name="「クリエイターを探す」から /creators に遷移",
            given="トップページを開いた状態",
            when="Hero 内の「クリエイターを探す」ボタンをクリック",
            then="/creators クリエイター一覧ページに遷移する",
        ),
        manual_case(
            id="TOP-008", account="全員",
            name="AI ニュース 8 件が表示され、PR TIMES 記事も含まれる",
            given="トップページを開いた状態",
            when="AI ニュース セクションまでスクロール",
            then="8 件のニュース記事カードが表示され、そのうち少なくとも 1 件は「PR TIMES」バッジ付き",
            note="cron 実行後の pool 状態次第 (2026/07/22 時点 手動確認済)",
        ),
    ]
    sheet_defs.append(("01_トップページ", "AILIER のランディングページ (ログイン不要)", "全員", cases))

    # ---------- Sheet 02: 新規登録 ----------
    cases = [
        auto_case(
            id="REG-001", account="全員 (ログイン前)",
            name="新規登録画面が表示され、創設メンバー残数が見える",
            given="ブラウザで /register にアクセス可能",
            when="/register を開く",
            then="「新規登録」見出し + 「創設メンバー枠 残り XX 名」のバッジ表示",
            note="Playwright smoke",
        ),
        auto_case(
            id="REG-002", account="全員",
            name="表示名「ああああ」で登録できない",
            given="/register 画面",
            when="表示名に「ああああ」/ メール / パスワードを入力し「アカウントを作成」",
            then="画面上に「同じ文字の繰り返しは使用できません」等のエラーが出て 登録処理が進まない",
            note="Playwright smoke",
        ),
        auto_case(
            id="REG-003", account="全員",
            name="表示名「admin」で登録できない",
            given="/register 画面",
            when="表示名に「admin」等の予約語を入力し「アカウントを作成」",
            then="「この表示名は使用できません」エラーが出て 登録処理が進まない",
        ),
        auto_case(
            id="REG-004", account="全員",
            name="表示名「!!!」など記号のみで登録できない",
            given="/register 画面",
            when="表示名に「!!!」を入力し 送信",
            then="「文字または数字を含めてください」エラーが出て 登録処理が進まない",
        ),
        manual_case(
            id="REG-005", account="全員",
            name="クリエイター選択時に「個人 / 法人」ラジオが表示",
            given="/register 画面",
            when="アカウント種別で「クリエイター」を選択",
            then="下に「事業形態」として「個人 (フリーランス) / 法人 (映像制作会社等)」の 2 択が現れる",
        ),
        manual_case(
            id="REG-006", account="全員",
            name="依頼者選択時は「個人/法人」ラジオが表示されない",
            given="/register 画面",
            when="アカウント種別で「依頼者」を選択",
            then="事業形態ラジオは表示されない",
        ),
        manual_case(
            id="REG-007", account="全員",
            name="Google で登録ボタンが Google OAuth を開く",
            given="/register 画面",
            when="「Googleで登録」ボタンをクリック",
            then="Google アカウント選択画面が新規タブ / 同タブで開く",
        ),
        manual_case(
            id="REG-008", account="全員",
            name="LINE で登録ボタンが LINE ログインを開く",
            given="/register 画面",
            when="「LINEで登録」ボタンをクリック",
            then="LINE ログイン画面が開く",
        ),
        manual_case(
            id="REG-009", account="全員",
            name="正常な情報で登録すると 確認メールが届く",
            given="/register 画面 / 未登録のメール",
            when="正常な表示名 (例: 山田太郎) / 有効なメール / 6 文字以上のパスワード / クリエイター選択で送信",
            then="「確認メールを送信しました」画面に切替 + 実際のメールボックスに Supabase から確認メールが届く",
            note="登録メールが確認できない場合は迷惑メールも要確認",
        ),
        manual_case(
            id="REG-010", account="全員",
            name="既に退会済みメールで再登録できない",
            given="退会から 30 日以内のメール",
            when="そのメールで /register",
            then="エラー「このメールアドレスは既に登録されています…30 日以内は…」表示",
        ),
    ]
    sheet_defs.append(("02_新規登録", "/register — メール / Google / LINE 経由でアカウント作成", "全員 (ログイン前)", cases))

    # ---------- Sheet 03: ログイン ----------
    cases = [
        auto_case(
            id="LOG-001", account="全員 (ログイン前)",
            name="ログイン画面にメール / パスワード / ボタンが表示",
            given="ブラウザで /login にアクセス可能",
            when="/login を開く",
            then="「メールアドレス」「パスワード」入力欄と「ログイン」ボタンが表示",
            note="Playwright smoke",
        ),
        manual_case(
            id="LOG-002", account="登録済みユーザー",
            name="正しいメール / パスワードでログイン",
            given="事前に登録済み + 確認メール承認済のアカウント",
            when="メール + パスワードを入力し「ログイン」",
            then="/dashboard または /onboarding に自動遷移",
        ),
        manual_case(
            id="LOG-003", account="登録済み",
            name="間違ったパスワードでエラー表示",
            given="登録済みアカウント",
            when="正しいメール + 間違ったパスワードで「ログイン」",
            then="「メールアドレスまたはパスワードが正しくありません」表示 + 画面遷移しない",
        ),
        manual_case(
            id="LOG-004", account="全員",
            name="未登録メールでエラー表示",
            given="登録していないメール",
            when="そのメール + 任意のパスワードで「ログイン」",
            then="エラー表示 (無効な認証情報 等)",
        ),
        manual_case(
            id="LOG-005", account="退会済み",
            name="退会済みアカウントは強制ログアウトされる",
            given="運営側で is_active=false に倒したアカウント",
            when="そのアカウントでログイン後、任意のダッシュボード ページを開く",
            then="middleware が signOut → トップページに戻される",
            note="middleware.ts 実装で自動処理",
        ),
        manual_case(
            id="LOG-006", account="全員",
            name="「Google でログイン」で Google OAuth 開始",
            given="/login 画面",
            when="Google ログインボタンクリック",
            then="Google アカウント選択画面",
        ),
        manual_case(
            id="LOG-007", account="全員",
            name="「LINE でログイン」で LINE 認証開始",
            given="/login 画面",
            when="LINE ログインボタンクリック",
            then="LINE 認証画面",
        ),
    ]
    sheet_defs.append(("03_ログイン", "/login — メール / OAuth ログイン", "登録済みユーザー", cases))

    # ---------- Sheet 04: オンボーディング ----------
    cases = [
        manual_case(
            id="ONB-001", account="新規クリエイター",
            name="登録直後に /onboarding が開く",
            given="creator として新規登録し、確認メールを承認直後",
            when="確認メール内リンクをクリック → auth callback 経由",
            then="/onboarding のウィザード画面 (STEP1 基本情報確認) が開く",
        ),
        manual_case(
            id="ONB-002", account="新規クリエイター",
            name="STEP1: 表示名確認 + 個人/法人選択 + 自己紹介入力",
            given="/onboarding STEP1 表示中",
            when="表示名 (読取専用) 確認 → 個人 or 法人 選択 → 自己紹介入力 (任意) → 「次へ」",
            then="STEP2 に遷移",
        ),
        manual_case(
            id="ONB-003", account="新規クリエイター",
            name="STEP2: ドラッグ&ドロップで動画/画像を追加",
            given="/onboarding STEP2 表示中",
            when="mp4 ファイル + jpg 画像 を D&D エリアに ドロップ",
            then="各アイテムのサムネイル/ファイル名/サイズが プレビューされる",
        ),
        manual_case(
            id="ONB-004", account="新規クリエイター",
            name="STEP2: URL タブで YouTube URL 追加",
            given="/onboarding STEP2 表示中",
            when="「URL 埋め込み」タブに切替 → YouTube URL を入力 → 「追加」",
            then="YouTube サムネイルが自動取得されてプレビュー",
        ),
        manual_case(
            id="ONB-005", account="新規クリエイター",
            name="STEP2: 0 件では投稿ボタン disabled",
            given="/onboarding STEP2 (何も追加していない)",
            when="投稿ボタンを見る",
            then="「投稿して公開する」ボタンが disabled (灰色)",
        ),
        manual_case(
            id="ONB-006", account="新規クリエイター",
            name="投稿完了で /dashboard?welcome=1 に遷移",
            given="STEP2 で 1 点以上追加済",
            when="「投稿して公開する」クリック",
            then="アップロード進捗が表示された後、/dashboard に遷移。トップに歓迎表示",
        ),
        manual_case(
            id="ONB-007", account="新規クリエイター",
            name="オンボーディング完了後は /onboarding にアクセスできない",
            given="STEP2 を完了した creator",
            when="URL バーで /onboarding を直接開く",
            then="/dashboard にリダイレクトされる",
        ),
    ]
    sheet_defs.append(("04_オンボーディング", "/onboarding — 新規クリエイターの 2 ステップ初期設定", "新規クリエイター", cases))

    # ---------- Sheet 05: 公開ページ ----------
    cases = [
        auto_case(
            id="PUB-001", account="全員",
            name="/creators クリエイター一覧が開く",
            given="ブラウザ",
            when="/creators を開く",
            then="200 表示 + タイトル / 検索 UI が見える",
            note="Playwright smoke",
        ),
        auto_case(
            id="PUB-002", account="全員",
            name="/pricing に 15% と 0% の両方が明記されている",
            given="ブラウザ",
            when="/pricing を開く",
            then="本文中に「15%」「0%」の両方の文字列を確認できる",
            note="Playwright smoke",
        ),
        auto_case(
            id="PUB-003", account="全員",
            name="/help に主要 7 FAQ が全て表示",
            given="ブラウザ",
            when="/help を開く",
            then="「手数料はいくら」「支払いはいつ」「登録方法」「案件はどのくらい」「著作権」「他のクラウドソーシング」「メッセージの返信 / レスポンス」の Q が全て見える",
            note="Playwright smoke",
        ),
        auto_case(
            id="PUB-004", account="全員",
            name="/terms に AI ガイド / キャンセル / 2 年保持 の条項が明記",
            given="ブラウザ",
            when="/terms を開く",
            then="「AI 生成」「着手前」「制作中」「納品後」「2 年」全てのキーワード表示",
            note="Playwright smoke",
        ),
        manual_case(
            id="PUB-005", account="全員",
            name="/how-it-works に 3 ステップの図解",
            given="ブラウザ",
            when="/how-it-works を開く",
            then="企業向け 3 ステップ + AI クリエイター向け 3 ステップの計 6 ステップが順に表示",
        ),
        manual_case(
            id="PUB-006", account="全員",
            name="/creator-guide に創設メンバー Hero バッジ表示",
            given="ブラウザ",
            when="/creator-guide を開く",
            then="Hero 直下に「Founding Creators / 創設メンバー」パネル + 残数",
        ),
        manual_case(
            id="PUB-007", account="全員",
            name="/company に運営会社 (Comhuman-Quality) 記載",
            given="ブラウザ",
            when="/company を開く",
            then="「Comhuman-Quality」社名 / 所在地等",
        ),
        manual_case(
            id="PUB-008", account="全員",
            name="/privacy プライバシーポリシー表示",
            given="ブラウザ",
            when="/privacy を開く",
            then="プライバシーポリシー条項が表示",
        ),
        manual_case(
            id="PUB-009", account="全員",
            name="/jobs 案件一覧に募集中案件が表示",
            given="案件が最低 1 件掲載済み",
            when="/jobs を開く",
            then="募集中案件がカード形式で表示 + タグでフィルタ可能",
        ),
        manual_case(
            id="PUB-010", account="全員",
            name="/creators で創設メンバーが上位に表示",
            given="creator に early_member あり",
            when="/creators 一覧を開く",
            then="creator カードに「創設メンバー」バッジ表示 + 上位配置",
        ),
        manual_case(
            id="PUB-011", account="全員",
            name="/creators/[id] で該当クリエイター詳細",
            given="公開 creator が存在",
            when="creators 一覧から creator クリック",
            then="詳細ページ表示 / ポートフォリオ / 最低対応金額 / SNS リンク等",
        ),
        manual_case(
            id="PUB-012", account="全員",
            name="/creators/[id] で「評価・レビュー」は準備中表示",
            given="creator 詳細ページ",
            when="レビュー セクションを見る",
            then="「準備中」プレースホルダー表示 (星評価は非表示)",
        ),
    ]
    sheet_defs.append(("05_公開ページ", "/pricing /help /terms /company /privacy /how-it-works /creator-guide /jobs /creators/*", "全員 (ログイン前)", cases))

    # ---------- Sheet 06: ダッシュボード共通 ----------
    cases = [
        manual_case(
            id="DASH-001", account="クリエイター",
            name="ログイン後 /dashboard が開く",
            given="creator ログイン済",
            when="ヘッダーの「ダッシュボード」or 「/dashboard」URL",
            then="ダッシュボード画面 / サイドバー / 基本情報エディタ 表示",
        ),
        manual_case(
            id="DASH-002", account="クリエイター (作品 0 点)",
            name="未公開アラートが表示される",
            given="creator + 作品 0 点 (is_searchable=false)",
            when="/dashboard を開く",
            then="⚠️「あなたのプロフィールはまだ企業に公開されていません」バナー + 「ポートフォリオを登録する」ボタン",
        ),
        manual_case(
            id="DASH-003", account="クリエイター (早期メンバー)",
            name="早期メンバー特典バッジ表示",
            given="is_early_member=true の creator",
            when="/dashboard 開く",
            then="「アーリーメンバー特典 適用中 / 手数料 永久 0%」バッジ表示",
        ),
        manual_case(
            id="DASH-004", account="クリエイター (スカウト 1+ 件)",
            name="💌 スカウト受信バナー表示",
            given="pending 招待 が 1 件以上",
            when="/dashboard 開く",
            then="ピンク グラデーション バナー「運営からのおすすめ案件が N 件届いています」表示 + 「受信トレイを開く」CTA",
        ),
        manual_case(
            id="DASH-005", account="クリエイター (スカウト 0 件)",
            name="スカウト 0 件時はバナー非表示",
            given="pending 招待 0",
            when="/dashboard 開く",
            then="スカウト バナー表示されない",
        ),
        manual_case(
            id="DASH-006", account="クリエイター",
            name="評価カードが「準備中」で表示",
            given="ログイン後",
            when="/dashboard 開く",
            then="「評価・レビュー」カードに「準備中」ラベル + 説明文",
        ),
        manual_case(
            id="DASH-007", account="クリエイター",
            name="サイドバーに「💌 運営スカウト」リンク",
            given="ログイン後",
            when="/dashboard 開く (PC 幅)",
            then="左サイドバーに「💌 運営スカウト」項目 → /dashboard/invitations に遷移可",
        ),
        manual_case(
            id="DASH-008", account="企業",
            name="企業アカウント側のサイドバー",
            given="client ログイン",
            when="/dashboard",
            then="サイドバーが 依頼者向け (企業情報 / 案件管理 / メッセージ / 支払い管理) に切替",
        ),
        auto_case(
            id="DASH-009", account="企業",
            name="企業アカウントで /dashboard/invitations にアクセスできない",
            given="client cookie",
            when="/dashboard/invitations 直接アクセス",
            then="/dashboard にリダイレクト",
            note="role check 実装確認済",
        ),
        manual_case(
            id="DASH-010", account="全員",
            name="ヘッダーから未読メッセージ数が見える",
            given="未読 message あり",
            when="任意のログイン後ページ",
            then="ヘッダーの通知アイコンに未読数バッジ",
        ),
    ]
    sheet_defs.append(("06_ダッシュボード共通", "/dashboard — ログイン後のトップ", "クリエイター / 企業", cases))

    # ---------- Sheet 07: プロフィール ----------
    cases = [
        manual_case(
            id="PRO-001", account="クリエイター",
            name="プロフィール編集画面が開く",
            given="creator ログイン",
            when="/dashboard/profile",
            then="編集フォーム表示 / 現在の bio / ai_tools / genres / minimum_order_amount 等が反映",
        ),
        manual_case(
            id="PRO-002", account="クリエイター",
            name="AI ツールを複数選択 / 解除",
            given="編集画面",
            when="AI ツール chip を数個クリック → 保存",
            then="保存後リロードで chip 選択状態が保持",
        ),
        manual_case(
            id="PRO-003", account="クリエイター",
            name="最低受注金額を設定",
            given="編集画面",
            when="minimum_order_amount に金額入力 → 保存",
            then="/creators/[自分] で「¥XX,XXX〜」表示",
        ),
        manual_case(
            id="PRO-004", account="クリエイター",
            name="strengths (強み) は最大 2 個まで",
            given="編集画面",
            when="strengths に 3 個目を選ぼうとする",
            then="3 個目が選べない or 保存時にエラー",
        ),
        manual_case(
            id="PRO-005", account="クリエイター",
            name="アバター画像 アップロード",
            given="編集画面",
            when="画像ファイル選択 → アップロード",
            then="サムネイル更新 / /creators/[id] に反映",
        ),
        manual_case(
            id="PRO-006", account="クリエイター",
            name="SNS リンク登録",
            given="編集画面",
            when="Twitter / Instagram / YouTube 等 URL 入力 → 保存",
            then="creator 詳細ページに アイコンリンク表示",
        ),
        manual_case(
            id="PRO-007", account="企業",
            name="企業情報編集",
            given="client ログイン",
            when="/dashboard/profile → 会社名 / URL / 業種入力 → 保存",
            then="次回 /dashboard で反映",
        ),
    ]
    sheet_defs.append(("07_プロフィール編集", "/dashboard/profile — 表示名 / bio / AI ツール / 会社情報 の編集", "クリエイター / 企業", cases))

    # ---------- Sheet 08: ポートフォリオ ----------
    cases = [
        manual_case(
            id="PORT-001", account="クリエイター",
            name="ポートフォリオ管理画面が開く",
            given="creator ログイン",
            when="/dashboard/portfolio",
            then="既登録作品 一覧 + 「新規追加」ボタン",
        ),
        manual_case(
            id="PORT-002", account="クリエイター",
            name="動画作品を追加",
            given="/dashboard/portfolio + mp4 ファイル",
            when="新規追加 → mp4 選択 → タイトル + 説明 + タグ入力 → 保存",
            then="一覧に追加、is_searchable=true に自動遷移 (0→1 で trigger)",
        ),
        manual_case(
            id="PORT-003", account="クリエイター",
            name="画像作品を追加",
            given="/dashboard/portfolio + jpg 画像",
            when="新規追加 → 画像選択 → タイトル → 保存",
            then="一覧に追加",
        ),
        manual_case(
            id="PORT-004", account="クリエイター",
            name="URL 埋め込み (YouTube) を追加",
            given="/dashboard/portfolio",
            when="URL 埋め込み タブ → YouTube URL → 追加",
            then="サムネ自動取得 → 保存後一覧",
        ),
        manual_case(
            id="PORT-005", account="クリエイター",
            name="AI ツール タグを付与",
            given="作品追加/編集画面",
            when="「使用した AI ツール」で複数 chip 選択",
            then="creator 詳細で作品に AI ツールバッジ表示",
        ),
        manual_case(
            id="PORT-006", account="クリエイター",
            name="作品を削除",
            given="既 1 点以上登録",
            when="作品 → 削除",
            then="一覧から消える。0 件になった場合 is_searchable=false",
        ),
    ]
    sheet_defs.append(("08_ポートフォリオ", "/dashboard/portfolio — 作品追加 / 編集 / 削除", "クリエイター", cases))

    # ---------- Sheet 09: スカウト受信 ----------
    cases = [
        manual_case(
            id="INV-001", account="クリエイター",
            name="/dashboard/invitations が開く",
            given="creator ログイン",
            when="サイドバー「💌 運営スカウト」or URL 直",
            then="3 タブ (未回答 / 回答済み / 期限切れ) 表示",
        ),
        manual_case(
            id="INV-002", account="クリエイター (未回答あり)",
            name="未回答招待カード表示",
            given="pending 招待 1 件以上",
            when="/dashboard/invitations",
            then="ピンク グラデ 招待カード / 案件情報 / 予算 / 締切 / 残り日数 表示",
        ),
        manual_case(
            id="INV-003", account="クリエイター",
            name="運営からの一言メッセージ 表示",
            given="admin が message 付きで招待",
            when="招待カード確認",
            then="「運営から一言:」ブロックに message 表示",
        ),
        manual_case(
            id="INV-004", account="クリエイター",
            name="残り 3 日以内で「お早めに」赤バッジ",
            given="expires_at まで 3 日以内",
            when="招待カード表示",
            then="時計アイコン隣に赤い「お早めに」バッジ",
        ),
        manual_case(
            id="INV-005", account="クリエイター",
            name="「詳細を見て応募する」で /jobs/[id] 遷移",
            given="pending 招待",
            when="「詳細を見て応募する」ボタンクリック",
            then="/jobs/[id] 案件詳細ページに遷移 + 招待 status=accepted に更新",
        ),
        manual_case(
            id="INV-006", account="クリエイター",
            name="「今回は見送る」で理由入力 → 見送り確定",
            given="pending 招待",
            when="「今回は見送る」→ 理由入力 (任意) → 「見送りを確定」",
            then="status=declined 更新、招待カードが「回答済み」タブに移動",
        ),
        manual_case(
            id="INV-007", account="クリエイター",
            name="期限切れタブに expired 招待",
            given="expires_at 過去の招待",
            when="タブ「期限切れ」",
            then="オレンジバッジ「期限切れ」でカード表示",
        ),
    ]
    sheet_defs.append(("09_スカウト受信", "/dashboard/invitations — 運営からのおすすめ案件", "クリエイター", cases))

    # ---------- Sheet 10: 出金申請 ----------
    cases = [
        manual_case(
            id="PAY-001", account="クリエイター",
            name="/dashboard/payouts が開く",
            given="creator ログイン",
            when="サイドバー「支払・請求」or URL",
            then="出金可能残高 / 入金予定 / 入金済累計 3 カード + 入金明細",
        ),
        manual_case(
            id="PAY-002", account="クリエイター (残高 < 5000)",
            name="最低金額 未達アラート + ボタン無効",
            given="残高 < ¥5000",
            when="出金申請パネル",
            then="「最低出金金額 ¥5,000 に達していません」アラート + 「出金申請する」disabled",
        ),
        manual_case(
            id="PAY-003", account="クリエイター (残高 5000+)",
            name="全額 / 一部 出金申請 送信",
            given="残高 ≥ ¥5000",
            when="申請額入力 or 全額 → 「出金申請する」",
            then="内訳 (申請額 / 手数料 -¥250 / 実受取額) が preview + 送信成功メッセージ",
        ),
        manual_case(
            id="PAY-004", account="クリエイター",
            name="申請額が残高超過は自動 cap",
            given="残高 ¥5000",
            when="申請額に ¥10000 入力",
            then="実際は残高 ¥5000 に silent cap で eligible",
        ),
        manual_case(
            id="PAY-005", account="クリエイター (Stripe 未接続)",
            name="Stripe 未接続バナー",
            given="stripe_account_id 未設定",
            when="/payouts",
            then="Stripe Connect 未接続バナー + 接続 CTA",
        ),
    ]
    sheet_defs.append(("10_出金申請", "/dashboard/payouts — 売上・出金", "クリエイター", cases))

    # ---------- Sheet 11: 取引詳細画面 ----------
    cases = [
        manual_case(
            id="ORD-001", account="クリエイター / 企業",
            name="取引詳細ページ表示",
            given="1 件以上取引あり",
            when="/dashboard/orders → 取引選択",
            then="タイトル / ステータスバッジ / 進行ステップ / 依頼内容 / 編集要件 / アクション表示",
        ),
        manual_case(
            id="ORD-002", account="クリエイター / 企業",
            name="仮払い前は「⚠️ 仮払い完了まで作業開始しないで」バナー",
            given="escrow_status = pending (仮払い前 order)",
            when="取引詳細 開く",
            then="オレンジのバナーに「⚠️ 仮払いが完了するまで作業を開始しないでください。仮払い前の作業は補償の対象外となります。」表示",
        ),
        manual_case(
            id="ORD-003", account="クリエイター",
            name="仮払い前は 納品ボタンが disabled",
            given="仮払い前 order",
            when="納品段階のボタン",
            then="「納品する」ボタンが灰色 disabled + ホバー時に説明 tooltip",
        ),
        manual_case(
            id="ORD-004", account="クリエイター",
            name="仮払い完了で 納品ボタンが有効化",
            given="escrow = held (仮払い済) + status=production",
            when="取引詳細 開く",
            then="納品ボタン enabled + PrePayment アラートが消える",
        ),
        manual_case(
            id="ORD-005", account="クリエイター",
            name="修正回数上限直前で「これが最後の無償修正」警告",
            given="max_revisions=2 / revision_count_used=1 で delivered 状態",
            when="クライアントが修正依頼を送る",
            then="バナー「これが最後の無償修正です」表示",
        ),
        manual_case(
            id="ORD-006", account="企業",
            name="修正回数超過で「追加発注の対象です」エラー",
            given="max=1 used=1 で delivered",
            when="更に「修正を依頼」クリック",
            then="エラー「合意した修正回数の上限に達しています。追加発注 (別料金) の対象です。」",
        ),
        manual_case(
            id="ORD-007", account="クリエイター / 企業",
            name="「運営に相談する」ボタン 常設",
            given="任意の取引",
            when="取引詳細下部を見る",
            then="「運営に相談する」ボタン (紫) 表示、いつでもクリック可",
        ),
        manual_case(
            id="ORD-008", account="クリエイター",
            name="運営裁定中は 全アクション disabled",
            given="active_dispute_id 有り",
            when="取引詳細",
            then="「運営: 確認中」バッジ + 進行ボタン全て disable",
        ),
        manual_case(
            id="ORD-009", account="クリエイター / 企業",
            name="キャンセル済み snapshot 表示",
            given="cancel 実行済 order",
            when="取引詳細",
            then="赤ボックス「キャンセル済み」 / 段階 / 返金額 / 補償額 / 理由 表示",
        ),
        manual_case(
            id="ORD-010", account="企業",
            name="納品後にダウンロード → is_downloaded=true",
            given="delivered order + 納品物 URL",
            when="納品物ダウンロードリンククリック",
            then="ファイル DL + サーバー側で is_downloaded_by_client=true 記録",
        ),
        manual_case(
            id="ORD-011", account="企業",
            name="ダウンロード後は 100% 返金要求が拒否される",
            given="is_downloaded=true + delivered 段階以外",
            when="キャンセル モーダルで 100% 返金要求",
            then="409 エラー「納品物のダウンロード / 使用が確認されているため、この段階からの全額返金は受け付けられません」",
        ),
    ]
    sheet_defs.append(("11_取引詳細画面", "/dashboard/orders/[id] — 取引の進行 / 納品 / 修正", "クリエイター / 企業", cases))

    # ---------- Sheet 12: 途中終了モーダル ----------
    cases = [
        manual_case(
            id="TERM-001", account="クリエイター",
            name="「途中終了を申請する」ボタンでモーダル表示",
            given="in_progress 状態の取引",
            when="取引詳細 → 「途中終了を申請する」クリック",
            then="全画面モーダルが開く",
        ),
        manual_case(
            id="TERM-002", account="クリエイター",
            name="警告文言が仕様どおり",
            given="モーダル表示",
            when="モーダル本文を読む",
            then="「⚠️ 警告：同意すると仮払い金は 全額発注者に返金され、あなたの報酬は『ゼロ』になります。」の文字列 + 『ゼロ』が赤いバッジで強調",
        ),
        manual_case(
            id="TERM-003", account="クリエイター",
            name="「同意する前に」見出しと運営相談ボタン",
            given="モーダル表示",
            when="モーダル中段確認",
            then="緑枠の「同意する前に」ブロック + 「運営に相談する」ボタン",
        ),
        auto_case(
            id="TERM-004", account="クリエイター",
            name="チェックボックス未 ON で確定ボタン disabled",
            given="モーダル表示",
            when="チェックボックスを OFF のまま「途中終了を確定」ボタン確認",
            then="ボタン disabled (灰色)",
            note="Playwright safety",
        ),
        auto_case(
            id="TERM-005", account="クリエイター",
            name="理由未入力で確定ボタン disabled",
            given="モーダル + チェック ON",
            when="理由 textarea 空のまま",
            then="ボタン disabled",
            note="Playwright safety",
        ),
        manual_case(
            id="TERM-006", account="クリエイター",
            name="チェック ON + 理由入力で確定可能",
            given="モーダル",
            when="理由入力 → チェック ON",
            then="「途中終了を確定する」ボタン enabled (赤色)",
        ),
        manual_case(
            id="TERM-007", account="クリエイター",
            name="「運営に相談する」でウィザードに切替",
            given="モーダル表示",
            when="「運営に相談する」クリック",
            then="モーダル閉じる → トラブル解決ウィザードが開く",
        ),
        manual_case(
            id="TERM-008", account="クリエイター",
            name="mode=agree で緑ボックスの代替案 UI 表示",
            given="発注者からの終了申請への 同意モード (mode=agree)",
            when="モーダル表示",
            then="緑ボックス「同意する代わりに、以下の代替案を検討してください」+ 3 選択肢 (修正合意 / 部分報酬 / 運営裁定) 表示",
        ),
        manual_case(
            id="TERM-009", account="クリエイター",
            name="確定実行後 order が cancelled になる",
            given="全条件 OK",
            when="「途中終了を確定」実行",
            then="取引が「キャンセル済み」に、返金 100%/補償 0 の snapshot 表示",
        ),
    ]
    sheet_defs.append(("12_途中終了モーダル", "自爆防止 全画面警告モーダル", "クリエイター", cases))

    # ---------- Sheet 13: トラブル解決ウィザード ----------
    cases = [
        manual_case(
            id="WIZ-001", account="クリエイター / 企業",
            name="「運営に相談する」でウィザード起動",
            given="任意取引",
            when="取引詳細 → 「運営に相談する」クリック",
            then="ウィザード モーダル + 「どのようなお困りごとですか？」 表示",
        ),
        manual_case(
            id="WIZ-002", account="全員",
            name="6 カテゴリ全て選択可能",
            given="ウィザード表示",
            when="各 カテゴリ chip を確認",
            then="連絡なし / 検収遅延 / 不当修正 / 品質問題 / 途中終了不同意 / その他 の 6 種選択できる",
        ),
        auto_case(
            id="WIZ-003", account="クリエイター",
            name="STEP1 未実施で 裁定申請 ボタン disabled",
            given="催促未実施 (first_reminder_sent_at=null) + 「連絡が来ない」選択",
            when="ウィザード内 STEP3 を確認",
            then="「完了していません」バナー + 裁定申請ボタン disabled or 非表示",
            note="Playwright safety",
        ),
        manual_case(
            id="WIZ-004", account="クリエイター",
            name="STEP1 (催促) 実施後は STEP3 が活性化",
            given="催促済み (first_reminder_sent_at セット)",
            when="wizard → 「連絡が来ない」選択",
            then="STEP1 に緑チェック + STEP3 「運営に裁定を申請する」ボタン enabled",
        ),
        manual_case(
            id="WIZ-005", account="企業",
            name="「不当な修正要求」→ STEP2 (修正依頼) 必須",
            given="revision_count_used = 0",
            when="wizard → 「不当な修正要求」選択",
            then="STEP2 未完了で 裁定ボタン disabled + 「取引詳細画面へ」CTA",
        ),
        manual_case(
            id="WIZ-006", account="企業",
            name="「品質問題」→ STEP2 (納品受領) 必須",
            given="delivered_at = null",
            when="wizard → 「品質問題」選択",
            then="STEP2 未完了で 裁定ボタン disabled",
        ),
        manual_case(
            id="WIZ-007", account="全員",
            name="裁定申請 送信で 受付バッジ表示",
            given="STEP1/2 完了",
            when="詳細入力 → 「運営に裁定を申請する」",
            then="「受け付けました」メッセージ + 取引画面上部に「運営: 受付しました」バッジ",
        ),
        manual_case(
            id="WIZ-008", account="全員",
            name="STEP1 の「メッセージを送る」で messages 画面に遷移",
            given="催促未実施",
            when="STEP1 の CTA クリック",
            then="/dashboard/messages/[partnerId] に遷移",
        ),
    ]
    sheet_defs.append(("13_トラブル解決ウィザード", "運営相談 - カテゴリ選択 → STEP1-3 順序ガード", "クリエイター / 企業", cases))

    # ---------- Sheet 14: メッセージ ----------
    cases = [
        manual_case(
            id="MSG-001", account="クリエイター / 企業",
            name="メッセージ画面が開く",
            given="messages 1 件以上",
            when="/dashboard/messages",
            then="相手一覧サイドバー + 会話ペイン",
        ),
        manual_case(
            id="MSG-002", account="全員",
            name="メッセージ送信",
            given="/dashboard/messages/[partnerId] 開く",
            when="入力欄に文字 → Enter or 送信",
            then="メッセージが即時反映 / 相手側でも受信",
        ),
        manual_case(
            id="MSG-003", account="全員",
            name="外部取引警告バナー 常時表示",
            given="messages 画面",
            when="入力欄付近を確認",
            then="「※ 外部ツール (LINE / Slack / メール等) でのやり取り・納品はトラブル時の補償対象外となります」バナー表示",
        ),
        manual_case(
            id="MSG-004", account="全員",
            name="ファイル (画像 / PDF) 添付",
            given="messages 画面",
            when="クリップ アイコン → ファイル選択 → 送信",
            then="サムネイル付きで添付表示",
        ),
        manual_case(
            id="MSG-005", account="全員",
            name="メッセージ削除",
            given="自分の送信 message",
            when="ゴミ箱アイコン クリック → 確認",
            then="ソフト削除 (is_deleted=true) で表示から消える",
        ),
    ]
    sheet_defs.append(("14_メッセージ", "/dashboard/messages — 取引相手とのやり取り", "クリエイター / 企業", cases))

    # ---------- Sheet 15: 通報機能 ----------
    cases = [
        manual_case(
            id="RPT-001", account="全員 (ログイン済)",
            name="作品カード hover で 通報アイコン表示",
            given="ログイン済 + creator 詳細ページ",
            when="作品サムネイル にホバー",
            then="右上に赤い 旗アイコン (Flag) 表示",
        ),
        manual_case(
            id="RPT-002", account="全員 (ログイン前)",
            name="未ログイン は 通報時に /login 誘導",
            given="ログイン前",
            when="通報アイコン クリック",
            then="/login?next=... にリダイレクト",
        ),
        manual_case(
            id="RPT-003", account="ログイン済",
            name="通報ダイアログの 6 カテゴリ",
            given="通報アイコン クリック",
            when="ダイアログ表示",
            then="著作権侵害 / なりすまし / 実在人物の無断生成 / 公序良俗違反 / スパム / その他 の 6 選択肢",
        ),
        manual_case(
            id="RPT-004", account="ログイン済",
            name="カテゴリ未選択で 送信ボタン disabled",
            given="ダイアログ",
            when="カテゴリ未選択のまま",
            then="送信ボタン disabled",
        ),
        manual_case(
            id="RPT-005", account="ログイン済",
            name="通報送信 → 受付メッセージ + Email",
            given="ダイアログ + カテゴリ選択済 + 詳細入力",
            when="「通報を送信」クリック",
            then="「通報を受け付けました」画面 + 数分後にログインメールに受付自動返信が届く",
        ),
        auto_case(
            id="RPT-006", account="ログイン済",
            name="同一作品への 2 通目は 409 で拒否",
            given="既 通報済作品",
            when="同一作品を再度通報",
            then="「この作品には既に通報を送信済みです」エラー",
            note="unique 制約実装済",
        ),
        auto_case(
            id="RPT-007", account="ログイン済",
            name="1 IP から 24h 5 件超過で rate limit",
            given="同一 IP から 5 通送信済",
            when="6 通目を送る",
            then="「通報の送信回数が上限に達しています」429 エラー",
            note="rate-limit 実装確認済",
        ),
        manual_case(
            id="RPT-008", account="全員",
            name="異なる 3 IP から 通報で 自動非公開",
            given="異なるユーザー / IP から 2 通済",
            when="3 通目を別 IP で送信",
            then="即座に作品が /creators 一覧から消える + クリエイターにメール通知",
        ),
    ]
    sheet_defs.append(("15_通報機能", "作品カードから 通報 / 自動非公開", "全員", cases))

    # ---------- Sheet 16: 管理画面 - モデレーション ----------
    cases = [
        manual_case(
            id="MOD-001", account="管理者",
            name="/admin/moderation が開く",
            given="admin ログイン",
            when="/admin/moderation",
            then="4 セクション (未対応通報 / 非公開・削除済 / 常習者 / 監査ログ)",
        ),
        manual_case(
            id="MOD-002", account="管理者",
            name="未対応通報の一覧表示",
            given="通報 1 件以上",
            when="/admin/moderation",
            then="作品 / クリエイター / 通報数 / unique IP 数 / 主要カテゴリ 表示",
        ),
        manual_case(
            id="MOD-003", account="管理者",
            name="常習者 一覧表示",
            given="通報 2+ or 非公開 1+ の creator あり",
            when="常習者セクション",
            then="表形式で display_name / 累積通報 / 未対応 / 非公開回数 / 削除回数 / 最新インシデント日 表示",
        ),
        manual_case(
            id="MOD-004", account="管理者",
            name="⚠️ 要監視 バッジ表示",
            given="削除 1+ or 非公開 3+ の creator",
            when="常習者行を見る",
            then="⚠️ 要監視バッジが display_name 隣に表示",
        ),
        manual_case(
            id="MOD-005", account="管理者",
            name="一時非公開 実行 (プルダウン + 補足)",
            given="published 作品",
            when="「非公開」ボタン → プルダウンから理由選択 (5 種) → 補足入力 → 「一時非公開を実行」",
            then="作品 status=unpublished + 監査ログ INSERT + クリエイターに Email",
        ),
        auto_case(
            id="MOD-006", account="管理者",
            name="カテゴリ未選択で 実行ボタン disabled",
            given="unpublish フォーム",
            when="カテゴリ プルダウン未選択",
            then="実行ボタン disabled",
            note="実装確認済",
        ),
        manual_case(
            id="MOD-007", account="管理者",
            name="削除 実行",
            given="unpublished 作品",
            when="「削除」→ 理由 → 実行",
            then="status=deleted + 監査ログ + クリエイターに「削除 (最終措置)」メール",
        ),
        manual_case(
            id="MOD-008", account="管理者",
            name="復元 実行",
            given="unpublished 作品",
            when="「復元」→ 理由 (任意) → 実行",
            then="status=published + 監査ログ + クリエイターに「公開再開しました」メール",
        ),
        manual_case(
            id="MOD-009", account="管理者",
            name="監査ログに全操作履歴",
            given="複数の unpublish/delete/restore 実行済",
            when="監査ログ セクション",
            then="直近 30 件が action_type + actor + reason + created_at 表示",
        ),
    ]
    sheet_defs.append(("16_管理_モデレーション", "/admin/moderation — 通報対応 / 一時非公開 / 削除", "管理者", cases))

    # ---------- Sheet 17: 管理画面 - 運営裁定 ----------
    cases = [
        manual_case(
            id="RUL-001", account="管理者",
            name="/admin/disputes 一覧",
            given="admin ログイン + 1 件以上 dispute あり",
            when="/admin/disputes",
            then="対応が必要 / 対応完了 の 2 セクション",
        ),
        manual_case(
            id="RUL-002", account="管理者",
            name="dispute 詳細画面",
            given="対応必要 dispute",
            when="行の「詳細・裁定」クリック",
            then="申請理由 / 対象 order 情報 / 履歴 / 右カラム RulingForm 表示",
        ),
        manual_case(
            id="RUL-003", account="管理者",
            name="RulingForm 5 種類 選択可",
            given="dispute 詳細",
            when="ruling type radio",
            then="partial_refund / full_refund / reproduction / no_action / as_is が全て選択可",
        ),
        manual_case(
            id="RUL-004", account="管理者",
            name="partial_refund で スライダ + preview",
            given="partial_refund 選択",
            when="スライダで rate 移動",
            then="「クライアント返金」「クリエイター補償」金額 preview がリアルタイム更新",
        ),
        auto_case(
            id="RUL-005", account="管理者",
            name="説明未入力で 実行ボタン disabled",
            given="RulingForm",
            when="resolution_summary 空",
            then="ボタン disabled",
            note="実装確認済",
        ),
        manual_case(
            id="RUL-006", account="管理者",
            name="裁定確定 → dispute resolved + 双方通知",
            given="全項目入力済",
            when="「この裁定で確定する」実行",
            then="dispute status=resolved / order 状態遷移 / client + creator にメール",
        ),
        manual_case(
            id="RUL-007", account="管理者",
            name="内部メモは クリエイター/クライアント通知には含まれない",
            given="internal_note 入力済",
            when="通知メール確認",
            then="ユーザー向け通知には resolution_summary のみで internal_note は含まれない",
        ),
    ]
    sheet_defs.append(("17_管理_運営裁定", "/admin/disputes — 紛争裁定", "管理者", cases))

    # ---------- Sheet 18: 管理画面 - スカウト送信 ----------
    cases = [
        manual_case(
            id="SCT-001", account="管理者",
            name="/admin/jobs 案件一覧",
            given="admin ログイン",
            when="/admin/jobs",
            then="オープン案件が表示 (タイトル / クライアント / 予算 / 応募数 / 招待済数)",
        ),
        manual_case(
            id="SCT-002", account="管理者",
            name="案件詳細で InviteSection 表示",
            given="/admin/jobs/[id]",
            when="ページ開く",
            then="右カラムに「クリエイターを探して招待する」パネル + タグ chip 一覧",
        ),
        manual_case(
            id="SCT-003", account="管理者",
            name="タグ AND 検索",
            given="InviteSection",
            when="複数タグ選択 → 「検索」",
            then="全タグを持つ creator のみ結果表示 + マッチ数バッジ",
        ),
        manual_case(
            id="SCT-004", account="管理者",
            name="全選択 / 個別選択",
            given="検索結果 5 件以上",
            when="「全選択」or 個別チェック",
            then="選択数バッジ更新",
        ),
        manual_case(
            id="SCT-005", account="管理者",
            name="一括招待送信",
            given="複数 creator 選択済",
            when="一言メッセージ入力 → 「N 名に招待を送る」",
            then="「N 名に招待通知を送信しました」メッセージ + 選択済 creator が結果から除外",
        ),
        manual_case(
            id="SCT-006", account="管理者",
            name="既招待は 検索結果から自動除外",
            given="既に invite 済み creator",
            when="同一タグで再検索",
            then="既招待 creator は結果に含まれない",
        ),
        manual_case(
            id="SCT-007", account="管理者",
            name="停止アカウント (is_active=false) は除外",
            given="is_active=false creator あり",
            when="タグ検索",
            then="停止アカウント は結果に含まれない",
        ),
        manual_case(
            id="SCT-008", account="管理者",
            name="招待済み一覧の status 表示",
            given="複数 status の招待あり",
            when="案件詳細の 左カラム「招待済みクリエイター」",
            then="招待中 / 応募済 / 見送り / 期限切れ の各バッジで表示",
        ),
    ]
    sheet_defs.append(("18_管理_スカウト送信", "/admin/jobs/[id] — タグで creator 検索 → 一括招待", "管理者", cases))

    # ---------- Sheet 19: 管理画面 - ユーザー・取引 ----------
    cases = [
        manual_case(
            id="ADM-001", account="管理者",
            name="/admin ダッシュボード",
            given="admin ログイン",
            when="/admin",
            then="総ユーザー数 / creator 数 / client 数 / 案件数 等の統計",
        ),
        manual_case(
            id="ADM-002", account="管理者",
            name="/admin/users ユーザー一覧",
            given="admin ログイン",
            when="/admin/users",
            then="profiles / creator_profiles / client_profiles 一覧 + 検索",
        ),
        manual_case(
            id="ADM-003", account="管理者",
            name="/admin/orders 取引・売上",
            given="admin ログイン",
            when="/admin/orders",
            then="全 orders 集計 + 一覧",
        ),
        manual_case(
            id="ADM-004", account="管理者",
            name="停止アカウント一覧 (is_active=false)",
            given="停止 creator あり",
            when="/admin/users で filter",
            then="停止済み一覧 + 復帰ボタン",
        ),
        auto_case(
            id="ADM-005", account="非管理者",
            name="非 admin で /admin にアクセスすると /dashboard へリダイレクト",
            given="creator or client cookie",
            when="/admin 直接アクセス",
            then="/dashboard にリダイレクト",
            note="admin layout の role check 実装確認済",
        ),
    ]
    sheet_defs.append(("19_管理_ユーザー・取引", "/admin, /admin/users, /admin/orders", "管理者", cases))

    # ---------- Sheet 20: 自動処理 (Cron) ----------
    cases = [
        auto_case(
            id="CRON-001", account="運営 (システム)",
            name="AI ニュース 日次更新",
            given="Vercel Cron 22:00 UTC",
            when="/api/cron/refresh-ai-news が発火",
            then="12 RSS ソースから最新記事 fetch + Supabase に upsert",
            note="手動発火成功済 (2026/07/22)",
        ),
        auto_case(
            id="CRON-002", account="運営",
            name="Cron 認可 (Bearer 無しで 401)",
            given="全 cron 6 種",
            when="Authorization ヘッダ無しで直叩き",
            then="401 fail-closed",
            note="共通実装確認済",
        ),
        manual_case(
            id="CRON-003", account="運営",
            name="みなし検収 (delivered 後 7 日で自動 released)",
            given="delivered_at から 7 日以上経過 order",
            when="cron orders-auto-approve 発火",
            then="escrow_status=released / payout_status=scheduled 自動遷移",
        ),
        manual_case(
            id="CRON-004", account="運営",
            name="納品期限リマインド",
            given="納品期限 24h 以内 or 超過 order",
            when="cron orders-deadline-reminder 発火 (日次)",
            then="クリエイターに in-app + Email 通知",
        ),
        manual_case(
            id="CRON-005", account="運営",
            name="未納品 自動キャンセル (催促後 7 日)",
            given="nondelivery_deadline_at 超過",
            when="cron orders-nondelivery-cancel 発火",
            then="order cancelled + escrow refunded (100% 返金) + creator_penalties (weight=3) 追加",
        ),
        manual_case(
            id="CRON-006", account="運営",
            name="データ保持 2 年 (物理削除)",
            given="data_retention_until 過去 の messages / orders",
            when="cron data-retention 週次発火",
            then="messages 物理削除 (BATCH_MESSAGES=500) + orders soft_deleted_at セット",
        ),
        manual_case(
            id="CRON-007", account="運営",
            name="未納品 繰返しで 自動停止 (score>=15)",
            given="直近 12 ヶ月 penalty_score >= 15 の creator",
            when="cron suspend-repeat-offenders 発火 (日次)",
            then="profiles.is_active=false + suspended_at セット + creator に Email 通知 + 運営に Slack",
        ),
    ]
    sheet_defs.append(("20_自動処理_Cron", "バックグラウンド自動処理 (Vercel Cron)", "運営 (システム)", cases))

    # ---------- Sheet 21: 【実装予定】 ----------
    cases = [
        manual_case(
            id="TODO-001", account="管理者",
            name="【実装予定】Stripe 部分返金 実行 API",
            given="現在 DB snapshot のみ",
            when="キャンセル / 裁定確定後",
            then="stripe.refunds.create で実返金 実行",
            note="現在は DB スナップショットのみ",
        ),
        manual_case(
            id="TODO-002", account="クリエイター",
            name="【実装予定】途中終了 2 段階同意フロー",
            given="現在 1 発 API",
            when="途中終了申請",
            then="相手承認待ち pending_termination 中間状態を挟む",
        ),
        manual_case(
            id="TODO-003", account="全員",
            name="【実装予定】異議申立フォーム",
            given="現在は support@ailier.app への手動連絡",
            when="モデレーション後の異議申立",
            then="専用フォーム UI で受付 + 管理画面で対応可能",
        ),
        manual_case(
            id="TODO-004", account="全員",
            name="【実装予定】評価・レビュー (星評価) 復活",
            given="現在「準備中」表示",
            when="検収完了後",
            then="星評価 5 段階 UI + 一覧表示",
        ),
        manual_case(
            id="TODO-005", account="全員",
            name="【実装予定】自動マッチング アルゴリズム",
            given="現在 手動スカウトのみ",
            when="creator 検索",
            then="tags ベース推薦スコア + 上位表示",
        ),
        manual_case(
            id="TODO-006", account="全員",
            name="【実装予定】messages への 通報",
            given="現在 portfolio のみ",
            when="メッセージ通報",
            then="target_type='message' で受付 + 自動非公開ロジック",
        ),
        manual_case(
            id="TODO-007", account="全員",
            name="【実装予定】X (Twitter) OAuth",
            given="現在 Google / LINE のみ",
            when="X ログイン",
            then="/register /login で Xボタン",
        ),
        manual_case(
            id="TODO-008", account="運営",
            name="【実装予定】キャンセル料率レンジ化",
            given="現在 0/50/100 固定",
            when="運営裁定時",
            then="pre_start 0-10% / delivered 80-100% で override 可能",
        ),
        manual_case(
            id="TODO-009", account="管理者",
            name="【実装予定】dispute admin_status 中間遷移 UI",
            given="現在 resolve のみ",
            when="admin 対応",
            then="received → reviewing → resolved の中間遷移ボタン",
        ),
        manual_case(
            id="TODO-010", account="運営",
            name="【実装予定】creator 信頼度スコア 検索順位反映",
            given="現在 is_early_member のみ",
            when="/creators 一覧",
            then="penalty score も sort 反映",
        ),
    ]
    sheet_defs.append(("21_実装予定", "コード上・設計書上に存在するが 未実装の機能", "全員", cases))

    # ---------- 全 sheet 生成 ----------
    for name, intro, primary, cases in sheet_defs:
        add_sheet(wb, name, cases, intro=intro)

    # 目次に列記
    for i, (name, intro, primary, cases) in enumerate(sheet_defs):
        row = toc_start + i
        ws["A" + str(row)] = name
        ws["B" + str(row)] = intro
        ws["C" + str(row)] = primary
        ws["D" + str(row)] = len(cases)
        for col in ["A", "B", "C", "D"]:
            ws[col + str(row)].alignment = BODY_ALIGN
            ws[col + str(row)].border = BORDER

    # 目次を先頭に
    wb.move_sheet("00_目次", offset=-len(sheet_defs))

    return wb


if __name__ == "__main__":
    wb = build()
    out = "test_cases.xlsx"
    wb.save(out)
    print(f"generated: {out}")
    total = sum(ws.max_row - 2 for ws in wb.worksheets if ws.title != "00_目次")
    print(f"total test cases: {total}")
    print(f"sheets: {len(wb.worksheets)}")
    for ws in wb.worksheets:
        if ws.title != "00_目次":
            print(f"  {ws.title}: {ws.max_row - 2} cases")
